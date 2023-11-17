#!/usr/bin/env python3
# file: gen-report.py
# date: Nov 16, 2023
# author: Jon David
# description:
#   Given an Excel file (*.xlsx), generate a report in the form of an ORG (*.org) text file.
#

import sys
import openpyxl

from reportGenerator import *


fname = "cs527-fa23-groups-20230905.xlsx"
name_ugrads_sheets = "midterm-undergrads"
name_grads_sheets = "midterm-grads"
selected_sheet_name = None  ## to be selected by command line args

## found the bug: column IDs are not sequential after Yaz separated the sheets...
## thanks Yaz for making it confusing <_< ...
rowID_list_ugrad = [3,7,8,11,13,15,16,17,19,20,21,23,24,25,27,28,29,30,32,34,35,36]
rowID_list_grad = [4,5,6,10,12,14,18,22,26,31,33]
selected_rowID_list = None  ## to be selected by command line args


wb = openpyxl.load_workbook( fname )
names_all_sheets = wb.get_sheet_names()
print( names_all_sheets )

if len( sys.argv ) < 2:
    print( STR_HELP )
    sys.exit(1)
    
class_selection = sys.argv[1]
selected_sheet_name = None
if class_selection == "grad":
    selected_sheet_name = name_grads_sheets
    selected_rowID_list = rowID_list_grad
elif class_selection == "ugrad":
    selected_sheet_name = name_ugrads_sheets
    selected_rowID_list = rowID_list_ugrad
else:
    print( STR_HELP )
    sys.exit(1)

sheet = None
for name in names_all_sheets:
    if name == selected_sheet_name:
        sheet = wb[name]
        break


numRows = sheet.max_row
numCols = sheet.max_column

print( "[debug] numRows: {}".format( numRows ))
print( "[debug] numCols: {}".format( numCols ))

factory = OrgReportFactory( sheet )

for rowID in selected_rowID_list:
    r = factory.genReport( rowID )  ## returns OrgReport
    #r.print_me()
    #print("[debug] rowID: {}, studentName: {}".format( rowID, r.studentName ))
    rec = r.studentName.split(',')
    outf = "{}.org".format( rec[0] ) ## lastname only## "test.org"
    r.write_to_file( outf )
    print( "[status] ({},{}) wrote file: {}".format( rowID, r.studentName, outf ))


